"""CORVAX - chart of accounts management.

GAP THIS CLOSES
    The chart of accounts was read-only. It was created once by the seeder with
    three levels and there was no endpoint - and no screen - to add, rename or
    retire an account. An accounting system that cannot extend its own chart is
    not usable for a real business, and a 4th or 5th level was impossible.

WHAT THIS PROVIDES
    * create an account under any parent, to any depth
    * rename / re-classify / activate / deactivate
    * a tree view that shows the hierarchy with balances
    * guards that protect the ledger

GUARDS (these are the accounting rules, not optional niceties)
    * A parent may not be postable. The moment an account gains a child it stops
      accepting entries, because a total must not also hold its own movements.
    * The child inherits account_type and statement_group from the parent, so a
      liability can never end up hanging under assets.
    * An account that already carries journal lines cannot change its type or be
      deleted - only deactivated - otherwise historical statements would shift.
    * Codes are unique per company and must sit under the parent's numeric range.
"""
from __future__ import annotations

from io import BytesIO
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.db import get_db
from app.dependencies import ensure_permission, get_current_user
from app.models import Account, JournalEntry, User
from app.models.finance import JournalLine
from app.services.audit import write_audit

router = APIRouter(prefix="/chart-of-accounts", tags=["chart of accounts"])

ACCOUNT_TYPES = ["ASSET", "LIABILITY", "EQUITY", "REVENUE", "EXPENSE"]
MAX_DEPTH = 8  # deep enough for any practical chart; guards against loops
MAX_IMPORT_BYTES = 10 * 1024 * 1024
IMPORT_COLUMNS = (
    "account_code", "name_ar", "name_en", "parent_code", "account_type",
    "statement_group", "is_cash", "active",
)


class AccountIn(BaseModel):
    company_id: int
    code: str = Field(min_length=1, max_length=30)
    name_ar: str = Field(min_length=2, max_length=250)
    name_en: str = Field(min_length=2, max_length=250)
    parent_code: str | None = None
    # Only needed for a root account; otherwise inherited from the parent.
    account_type: str | None = None
    statement_group: str | None = None
    is_cash: bool = False


class AccountUpdate(BaseModel):
    company_id: int
    name_ar: str | None = Field(default=None, min_length=2, max_length=250)
    name_en: str | None = Field(default=None, min_length=2, max_length=250)
    is_cash: bool | None = None
    active: bool | None = None


def _get(db: Session, company_id: int, code: str) -> Account:
    account = db.scalar(select(Account).where(Account.company_id == company_id, Account.code == code))
    if not account:
        raise HTTPException(404, f"Account not found: {code}")
    return account


def _has_movements(db: Session, account_id: int) -> int:
    return db.scalar(select(func.count(JournalLine.id)).where(JournalLine.account_id == account_id)) or 0


def _bool(value, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().casefold()
    if normalized in {"true", "1", "yes", "y", "نعم"}:
        return True
    if normalized in {"false", "0", "no", "n", "لا"}:
        return False
    raise ValueError(f"Invalid boolean value: {value}")


def _excel_bytes(workbook: Workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _excel_safe(value):
    """Keep exported user-entered text from becoming an Excel formula."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _parse_import(file: UploadFile) -> list[dict]:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(422, "Only .xlsx files are accepted")
    content = file.file.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(413, "Workbook exceeds the 10 MB import limit")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise HTTPException(422, "The Excel workbook is invalid or unreadable") from exc
    sheet = workbook["Chart_of_Accounts"] if "Chart_of_Accounts" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise HTTPException(422, "The workbook is empty") from exc
    headers = [str(value or "").strip().casefold() for value in raw_headers]
    missing = [column for column in IMPORT_COLUMNS if column not in headers]
    if missing:
        raise HTTPException(422, {"message_ar": f"أعمدة مفقودة: {', '.join(missing)}", "message_en": f"Missing columns: {', '.join(missing)}"})
    index = {name: headers.index(name) for name in IMPORT_COLUMNS}
    rows: list[dict] = []
    for excel_row, values in enumerate(iterator, start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        def cell(name: str):
            position = index[name]
            return values[position] if position < len(values) else None
        if any(isinstance(cell(name), str) and cell(name).lstrip().startswith(("=", "+", "-", "@")) for name in IMPORT_COLUMNS):
            raise HTTPException(422, f"Excel formulas are not allowed in import row {excel_row}")
        rows.append({
            "excel_row": excel_row,
            "account_code": str(cell("account_code") or "").strip(),
            "name_ar": str(cell("name_ar") or "").strip(),
            "name_en": str(cell("name_en") or "").strip(),
            "parent_code": str(cell("parent_code") or "").strip() or None,
            "account_type": str(cell("account_type") or "").strip().upper(),
            "statement_group": str(cell("statement_group") or "").strip(),
            "is_cash_raw": cell("is_cash"),
            "active_raw": cell("active"),
        })
    if not rows:
        raise HTTPException(422, "No account rows were found")
    return rows


def _validate_import(db: Session, company_id: int, source_rows: list[dict]) -> dict:
    existing_rows = db.scalars(select(Account).where(Account.company_id == company_id)).all()
    existing = {row.code: row for row in existing_rows}
    staged: dict[str, dict] = {}
    results: list[dict] = []

    for source in source_rows:
        errors: list[str] = []
        warnings: list[str] = []
        code = source["account_code"]
        if not code:
            errors.append("ACCOUNT_CODE_REQUIRED")
        if code in staged:
            errors.append("DUPLICATE_ACCOUNT_CODE")
        if not source["name_ar"] or not source["name_en"]:
            errors.append("BILINGUAL_NAMES_REQUIRED")
        try:
            is_cash = _bool(source["is_cash_raw"], False)
            active = _bool(source["active_raw"], True)
        except ValueError as exc:
            errors.append(str(exc))
            is_cash, active = False, True
        current = existing.get(code)
        action = "UPDATE" if current else "CREATE"
        if current:
            current_parent = db.get(Account, current.parent_id).code if current.parent_id else None
            if source["parent_code"] != current_parent:
                errors.append("PARENT_CHANGE_NOT_ALLOWED")
            if source["account_type"] and source["account_type"] != current.account_type:
                errors.append("ACCOUNT_TYPE_CHANGE_NOT_ALLOWED")
            if source["statement_group"] and source["statement_group"] != current.statement_group:
                errors.append("STATEMENT_GROUP_CHANGE_NOT_ALLOWED")
            if not active:
                children = db.scalar(select(func.count(Account.id)).where(Account.parent_id == current.id, Account.active.is_(True))) or 0
                if children:
                    errors.append("ACTIVE_CHILDREN_PREVENT_DEACTIVATION")
        staged[code] = {
            **source,
            "is_cash": is_cash,
            "active": active,
            "action": action,
            "errors": errors,
            "warnings": warnings,
        }

    def candidate(code: str) -> dict | None:
        if code in staged:
            return staged[code]
        row = existing.get(code)
        if not row:
            return None
        parent = db.get(Account, row.parent_id).code if row.parent_id else None
        return {
            "account_code": row.code, "parent_code": parent, "account_type": row.account_type,
            "statement_group": row.statement_group, "active": row.active, "errors": [],
        }

    for code, row in staged.items():
        parent_code = row["parent_code"]
        if parent_code:
            parent = candidate(parent_code)
            if not parent:
                row["errors"].append("PARENT_ACCOUNT_NOT_FOUND")
                continue
            if parent_code == code:
                row["errors"].append("ACCOUNT_CANNOT_PARENT_ITSELF")
                continue
            visited = {code}
            cursor = parent
            depth = 1
            while cursor and cursor.get("parent_code"):
                cursor_code = cursor["parent_code"]
                if cursor_code in visited:
                    row["errors"].append("ACCOUNT_HIERARCHY_CYCLE")
                    break
                visited.add(cursor_code)
                cursor = candidate(cursor_code)
                depth += 1
            if depth + 1 > MAX_DEPTH:
                row["errors"].append("MAXIMUM_DEPTH_EXCEEDED")
            prefix = parent_code.rstrip("0") or parent_code[:1]
            if not code.startswith(prefix):
                row["errors"].append(f"CODE_OUTSIDE_PARENT_RANGE:{prefix}")
            if row["account_type"] and row["account_type"] != parent["account_type"]:
                row["errors"].append("CHILD_ACCOUNT_TYPE_MUST_MATCH_PARENT")
            row["account_type"] = parent["account_type"]
            row["statement_group"] = parent["statement_group"]
            existing_parent = existing.get(parent_code)
            if row["action"] == "CREATE" and existing_parent and existing_parent.is_postable and _has_movements(db, existing_parent.id):
                row["errors"].append("PARENT_HAS_POSTED_MOVEMENTS")
        else:
            if row["account_type"] not in ACCOUNT_TYPES:
                row["errors"].append("ROOT_ACCOUNT_TYPE_REQUIRED")
            if not row["statement_group"]:
                row["statement_group"] = row["account_type"]

    for row in staged.values():
        results.append({
            "excel_row": row["excel_row"],
            "account_code": row["account_code"],
            "action": row["action"],
            "errors": list(dict.fromkeys(row["errors"])),
            "warnings": list(dict.fromkeys(row["warnings"])),
        })
    error_count = sum(len(row["errors"]) for row in staged.values())
    return {
        "valid": error_count == 0,
        "rows": results,
        "summary": {
            "total_rows": len(staged),
            "create": sum(1 for row in staged.values() if row["action"] == "CREATE"),
            "update": sum(1 for row in staged.values() if row["action"] == "UPDATE"),
            "errors": error_count,
        },
        "_staged": staged,
    }


@router.get("/export.xlsx")
def export_chart(company_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "finance.read")
    rows = db.scalars(select(Account).where(Account.company_id == company_id).order_by(Account.code)).all()
    parent_codes = {row.id: row.code for row in rows}
    balances = dict(db.execute(
        select(
            JournalLine.account_id,
            func.coalesce(func.sum(JournalLine.debit - JournalLine.credit), 0),
        )
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_id)
        .where(JournalEntry.company_id == company_id, JournalEntry.status == "POSTED")
        .group_by(JournalLine.account_id)
    ).all())
    movements = dict(db.execute(
        select(JournalLine.account_id, func.count(JournalLine.id))
        .join(JournalEntry, JournalEntry.id == JournalLine.journal_id)
        .where(JournalEntry.company_id == company_id, JournalEntry.status == "POSTED")
        .group_by(JournalLine.account_id)
    ).all())
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Chart_of_Accounts"
    headers = [
        *IMPORT_COLUMNS, "level", "is_postable", "movement_lines", "current_balance",
        "import_action", "validation_status",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            _excel_safe(row.code), _excel_safe(row.name_ar), _excel_safe(row.name_en),
            _excel_safe(parent_codes.get(row.parent_id)), row.account_type,
            _excel_safe(row.statement_group), row.is_cash, row.active, row.level, row.is_postable,
            int(movements.get(row.id, 0)), Decimal(balances.get(row.id, 0)), "", "",
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.rightToLeft = True
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = {
        "A": 18, "B": 34, "C": 34, "D": 18, "E": 18, "F": 22,
        "G": 12, "H": 12, "I": 10, "J": 14, "K": 15, "L": 18, "M": 16, "N": 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    output = _excel_bytes(workbook)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="CORVAX_COA_{company_id}.xlsx"'},
    )


@router.post("/import/validate")
def validate_chart_import(
    company_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_permission(db, user, company_id, "masterdata.import")
    result = _validate_import(db, company_id, _parse_import(file))
    result.pop("_staged", None)
    return result


@router.post("/import/apply")
def apply_chart_import(
    company_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_permission(db, user, company_id, "masterdata.import")
    result = _validate_import(db, company_id, _parse_import(file))
    if not result["valid"]:
        result.pop("_staged", None)
        raise HTTPException(422, result)
    staged: dict[str, dict] = result["_staged"]
    existing = {row.code: row for row in db.scalars(select(Account).where(Account.company_id == company_id)).all()}

    for code, row in staged.items():
        current = existing.get(code)
        if current:
            before = {
                "name_ar": current.name_ar, "name_en": current.name_en,
                "is_cash": current.is_cash, "active": current.active,
            }
            current.name_ar = row["name_ar"]
            current.name_en = row["name_en"]
            current.is_cash = row["is_cash"]
            current.active = row["active"]
            write_audit(
                db, action="ACCOUNT_IMPORTED_UPDATE", entity_type="ACCOUNT", entity_id=current.id,
                user_id=user.id, company_id=company_id, before=before,
                after={"name_ar": current.name_ar, "name_en": current.name_en, "is_cash": current.is_cash, "active": current.active},
            )

    pending = {code for code, row in staged.items() if row["action"] == "CREATE"}
    while pending:
        progressed = False
        for code in sorted(tuple(pending)):
            row = staged[code]
            parent = existing.get(row["parent_code"]) if row["parent_code"] else None
            if row["parent_code"] and parent is None:
                continue
            account = Account(
                company_id=company_id, code=code, name_ar=row["name_ar"], name_en=row["name_en"],
                account_type=parent.account_type if parent else row["account_type"],
                statement_group=parent.statement_group if parent else row["statement_group"],
                parent_id=parent.id if parent else None,
                level=int(parent.level) + 1 if parent else 1,
                is_postable=True, is_cash=row["is_cash"], active=row["active"],
            )
            db.add(account)
            db.flush()
            if parent is not None:
                parent.is_postable = False
            existing[code] = account
            pending.remove(code)
            progressed = True
            write_audit(
                db, action="ACCOUNT_IMPORTED_CREATE", entity_type="ACCOUNT", entity_id=account.id,
                user_id=user.id, company_id=company_id,
                after={"code": code, "parent": row["parent_code"], "level": account.level},
            )
        if not progressed:
            db.rollback()
            raise HTTPException(422, "Account hierarchy could not be resolved")
    db.commit()
    result.pop("_staged", None)
    return {**result, "applied": True}


@router.get("/tree")
def account_tree(company_id: int, include_inactive: bool = False,
                 user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """The full chart as a nested tree, with per-account movement counts.

    The UI uses this to render levels 1..N and to show which accounts may still
    receive entries.
    """
    ensure_permission(db, user, company_id, "finance.read")
    query = select(Account).where(Account.company_id == company_id)
    if not include_inactive:
        query = query.where(Account.active.is_(True))
    rows = db.scalars(query.order_by(Account.code)).all()

    counts = dict(
        db.execute(
            select(JournalLine.account_id, func.count(JournalLine.id)).group_by(JournalLine.account_id)
        ).all()
    )

    by_id: dict[int, dict] = {}
    for row in rows:
        by_id[row.id] = {
            "id": row.id,
            "code": row.code,
            "name_ar": row.name_ar,
            "name_en": row.name_en,
            "account_type": row.account_type,
            "statement_group": row.statement_group,
            "level": row.level,
            "is_postable": row.is_postable,
            "is_cash": row.is_cash,
            "active": row.active,
            "parent_id": row.parent_id,
            "movement_lines": int(counts.get(row.id, 0)),
            "children": [],
        }

    roots: list[dict] = []
    for node in by_id.values():
        parent = by_id.get(node["parent_id"]) if node["parent_id"] else None
        (parent["children"] if parent else roots).append(node)

    def depth(nodes: list[dict], current: int = 1) -> int:
        return max([depth(n["children"], current + 1) for n in nodes if n["children"]] + [current])

    return {
        "company_id": company_id,
        "total_accounts": len(rows),
        "max_level": depth(roots) if roots else 0,
        "postable_accounts": sum(1 for r in rows if r.is_postable),
        "tree": sorted(roots, key=lambda n: n["code"]),
    }


@router.post("", status_code=201)
def create_account(data: AccountIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Add an account at any depth under an existing parent."""
    ensure_permission(db, user, data.company_id, "finance.manage")

    if db.scalar(select(Account).where(Account.company_id == data.company_id, Account.code == data.code)):
        raise HTTPException(409, f"Account code already exists: {data.code}")

    parent: Account | None = None
    level = 1
    account_type = (data.account_type or "").upper()
    statement_group = data.statement_group or ""

    if data.parent_code:
        parent = _get(db, data.company_id, data.parent_code)
        level = int(parent.level) + 1
        if level > MAX_DEPTH:
            raise HTTPException(422, f"Maximum depth of {MAX_DEPTH} levels reached")
        # A child always belongs to the same statement side as its parent.
        account_type = parent.account_type
        statement_group = parent.statement_group
        # The child must live inside the parent's numeric range. The parent's
        # significant prefix is its code without the trailing zeros:
        #   600000 -> "6"      so 630000 is valid
        #   613010 -> "61301"  so 613011 is valid
        prefix = data.parent_code.rstrip("0") or data.parent_code[:1]
        if not data.code.startswith(prefix):
            raise HTTPException(
                422,
                {
                    "message_ar": (
                        f"رقم الحساب يجب أن يبدأ بـ{prefix} ليقع داخل نطاق الحساب الأب {data.parent_code}. "
                        f"مثال صحيح: {prefix}{'1'.zfill(max(1, len(data.parent_code) - len(prefix)))}"
                    ),
                    "message_en": (
                        f"The code must start with {prefix} to sit inside parent {data.parent_code}."
                    ),
                },
            )
    else:
        if account_type not in ACCOUNT_TYPES:
            raise HTTPException(422, f"account_type is required for a root account. One of {ACCOUNT_TYPES}")
        if not statement_group:
            statement_group = account_type

    account = Account(
        company_id=data.company_id,
        code=data.code,
        name_ar=data.name_ar,
        name_en=data.name_en,
        account_type=account_type,
        statement_group=statement_group,
        parent_id=parent.id if parent else None,
        level=level,
        is_postable=True,          # a new leaf accepts entries
        is_cash=data.is_cash,
        active=True,
    )
    db.add(account)

    # A parent must stop accepting entries: a total cannot also hold movements.
    demoted = None
    if parent is not None and parent.is_postable:
        if _has_movements(db, parent.id):
            raise HTTPException(
                422,
                {
                    "message_ar": (
                        f"الحساب {parent.code} عليه حركات مرحّلة، فلا يمكن جعله حسابًا رئيسيًا. "
                        "انقل حركاته أولًا أو أنشئ الحساب تحت أب آخر."
                    ),
                    "message_en": (
                        f"Account {parent.code} already carries posted lines, so it cannot become a "
                        "parent. Move its entries first, or choose another parent."
                    ),
                },
            )
        parent.is_postable = False
        demoted = parent.code

    db.flush()
    write_audit(db, action="ACCOUNT_CREATED", entity_type="ACCOUNT", entity_id=account.id,
                user_id=user.id, company_id=data.company_id,
                after={"code": account.code, "level": level, "parent": data.parent_code})
    db.commit()

    return {
        "id": account.id,
        "code": account.code,
        "name_ar": account.name_ar,
        "level": level,
        "account_type": account_type,
        "statement_group": statement_group,
        "is_postable": True,
        "parent_demoted_to_header": demoted,
        "message_ar": (
            f"تم إنشاء الحساب في المستوى {level}."
            + (f" وتحوّل الحساب {demoted} إلى حساب رئيسي لا يقبل الترحيل." if demoted else "")
        ),
        "message_en": (
            f"Account created at level {level}."
            + (f" {demoted} became a header account and no longer accepts entries." if demoted else "")
        ),
    }


@router.patch("/{code}")
def update_account(code: str, data: AccountUpdate,
                   user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Rename, flag as cash, or retire an account. Type is never changed here."""
    ensure_permission(db, user, data.company_id, "finance.manage")
    account = _get(db, data.company_id, code)
    before = {"name_ar": account.name_ar, "active": account.active}

    if data.name_ar is not None:
        account.name_ar = data.name_ar
    if data.name_en is not None:
        account.name_en = data.name_en
    if data.is_cash is not None:
        account.is_cash = data.is_cash
    if data.active is not None:
        if not data.active:
            children = db.scalar(
                select(func.count(Account.id)).where(
                    Account.parent_id == account.id, Account.active.is_(True)
                )
            ) or 0
            if children:
                raise HTTPException(
                    422,
                    {
                        "message_ar": f"لا يمكن تعطيل {code} لأن تحته {children} حسابًا نشطًا. عطّلها أولًا.",
                        "message_en": f"Cannot deactivate {code}: {children} active child accounts remain.",
                    },
                )
        account.active = data.active

    write_audit(db, action="ACCOUNT_UPDATED", entity_type="ACCOUNT", entity_id=account.id,
                user_id=user.id, company_id=data.company_id, before=before,
                after={"name_ar": account.name_ar, "active": account.active})
    db.commit()
    return {"code": account.code, "name_ar": account.name_ar, "active": account.active,
            "is_postable": account.is_postable, "level": account.level}


@router.delete("/{code}")
def delete_account(code: str, company_id: int,
                   user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Remove an account that was never used. Used accounts are deactivated instead."""
    ensure_permission(db, user, company_id, "finance.manage")
    account = _get(db, company_id, code)

    movements = _has_movements(db, account.id)
    if movements:
        raise HTTPException(
            422,
            {
                "message_ar": f"الحساب {code} عليه {movements} سطر قيد، فلا يُحذف. عطّله بدلًا من ذلك.",
                "message_en": f"{code} carries {movements} journal lines and cannot be deleted. Deactivate it instead.",
            },
        )
    children = db.scalar(select(func.count(Account.id)).where(Account.parent_id == account.id)) or 0
    if children:
        raise HTTPException(422, f"{code} has {children} child accounts. Remove them first.")

    parent_id = account.parent_id
    write_audit(db, action="ACCOUNT_DELETED", entity_type="ACCOUNT", entity_id=account.id,
                user_id=user.id, company_id=company_id, before={"code": code})
    db.delete(account)
    db.flush()

    # If the parent has no children left it can accept entries again.
    restored = None
    if parent_id:
        remaining = db.scalar(select(func.count(Account.id)).where(Account.parent_id == parent_id)) or 0
        if remaining == 0:
            parent = db.get(Account, parent_id)
            if parent is not None:
                parent.is_postable = True
                restored = parent.code
    db.commit()
    return {"deleted": code, "parent_restored_to_postable": restored}
