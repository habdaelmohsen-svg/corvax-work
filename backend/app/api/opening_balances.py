from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.core.time import utc_now
from app.db import get_db
from app.dependencies import ensure_permission, get_current_user
from app.models import (
    Account, FinancialOpenItem, Item, JournalEntry, OpeningBalanceBatch, OpeningBalanceLine,
    Party, StockMovement, User, Warehouse,
)
from app.services.audit import write_audit
from app.services.posting import create_posted_journal, ensure_open_period


router = APIRouter(prefix="/opening-balances", tags=["opening balances"])
MONEY = Decimal("0.01")
FOUR = Decimal("0.0001")
MAX_IMPORT_BYTES = 15 * 1024 * 1024
LINE_TYPES = {"GL", "AR", "AP", "INVENTORY"}
HEADERS = (
    "line_type", "account_code", "party_code", "item_code", "warehouse_code",
    "reference_code", "document_date", "due_date", "quantity", "unit_cost",
    "lot_number", "debit", "credit", "description",
)


def money(value) -> Decimal:
    try:
        return Decimal(str(value or 0)).quantize(MONEY, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"INVALID_AMOUNT:{value}") from exc


def quantity(value) -> Decimal:
    try:
        return Decimal(str(value or 0)).quantize(FOUR, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise ValueError(f"INVALID_QUANTITY:{value}") from exc


def parse_date(value, *, required: bool = False) -> date | None:
    if value is None or value == "":
        if required:
            raise ValueError("DATE_REQUIRED")
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip()[:10])
    except ValueError as exc:
        raise ValueError(f"INVALID_DATE:{value}") from exc


def _hash_rows(rows: list[dict]) -> str:
    canonical = json.dumps(
        [
            {
                key: (
                    value.isoformat()
                    if isinstance(value, date)
                    else format(value.normalize(), "f")
                    if isinstance(value, Decimal)
                    else value
                )
                for key, value in row.items()
                if not key.startswith("_")
            }
            for row in rows
        ],
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _excel_safe(value):
    """Keep exported master-data text from becoming an Excel formula."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _stored_hash(batch: OpeningBalanceBatch) -> str:
    return _hash_rows([
        {
            "line_number": line.line_number,
            "line_type": line.line_type,
            "account_code": line.account.code,
            "party_code": line.party.code if line.party else None,
            "item_code": line.item.code if line.item else None,
            "warehouse_code": line.warehouse.code if line.warehouse else None,
            "reference_code": line.reference_code,
            "document_date": line.document_date,
            "due_date": line.due_date,
            "quantity": Decimal(line.quantity or 0),
            "unit_cost": Decimal(line.unit_cost or 0),
            "lot_number": line.lot_number,
            "debit": Decimal(line.debit),
            "credit": Decimal(line.credit),
            "description": line.description,
        }
        for line in batch.lines
    ])


def _parse_workbook(file: UploadFile) -> list[dict]:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(422, "Only .xlsx files are accepted")
    content = file.file.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(413, "Workbook exceeds the 15 MB import limit")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise HTTPException(422, "The Excel workbook is invalid or unreadable") from exc
    sheet = workbook["Opening_Balances"] if "Opening_Balances" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise HTTPException(422, "The workbook is empty") from exc
    headers = [str(value or "").strip().casefold() for value in raw_headers]
    missing = [column for column in ("line_type", "account_code", "debit", "credit") if column not in headers]
    if missing:
        raise HTTPException(422, {"message_ar": f"أعمدة مفقودة: {', '.join(missing)}", "message_en": f"Missing columns: {', '.join(missing)}"})
    indexes = {name: headers.index(name) for name in HEADERS if name in headers}
    rows: list[dict] = []
    for excel_row, values in enumerate(iterator, start=2):
        def cell(name: str):
            position = indexes.get(name)
            return values[position] if position is not None and position < len(values) else None
        if all(cell(name) is None or str(cell(name)).strip() == "" for name in indexes):
            continue
        if any(isinstance(cell(name), str) and cell(name).lstrip().startswith(("=", "+", "-", "@")) for name in indexes):
            raise HTTPException(422, f"Excel formulas are not allowed in import row {excel_row}")
        try:
            debit, credit = money(cell("debit")), money(cell("credit"))
        except ValueError:
            debit, credit = Decimal("0"), Decimal("0")
        if debit == 0 and credit == 0 and not str(cell("account_code") or "").strip():
            continue
        rows.append({
            "line_number": excel_row,
            "line_type": str(cell("line_type") or "GL").strip().upper(),
            "account_code": str(cell("account_code") or "").strip(),
            "party_code": str(cell("party_code") or "").strip() or None,
            "item_code": str(cell("item_code") or "").strip() or None,
            "warehouse_code": str(cell("warehouse_code") or "").strip() or None,
            "reference_code": str(cell("reference_code") or "").strip() or None,
            "document_date_raw": cell("document_date"),
            "due_date_raw": cell("due_date"),
            "quantity_raw": cell("quantity"),
            "unit_cost_raw": cell("unit_cost"),
            "lot_number": str(cell("lot_number") or "").strip() or None,
            "debit_raw": cell("debit"),
            "credit_raw": cell("credit"),
            "description": str(cell("description") or "").strip() or None,
        })
    if not rows:
        raise HTTPException(422, "No opening-balance rows were found")
    return rows


def _validate(db: Session, company_id: int, source_rows: list[dict]) -> dict:
    accounts = {row.code: row for row in db.scalars(select(Account).where(Account.company_id == company_id)).all()}
    parties = {row.code: row for row in db.scalars(select(Party).where(Party.company_id == company_id)).all()}
    items = {row.code: row for row in db.scalars(select(Item).where(Item.company_id == company_id)).all()}
    warehouses = {row.code: row for row in db.scalars(select(Warehouse).where(Warehouse.company_id == company_id)).all()}
    inventory_account_ids = {row.inventory_account_id for row in items.values()}
    existing_documents = set(db.scalars(select(FinancialOpenItem.document_number).where(FinancialOpenItem.company_id == company_id)).all())
    document_keys: set[tuple[str, str]] = set()
    prepared: list[dict] = []
    total_debit = Decimal("0")
    total_credit = Decimal("0")

    for source in source_rows:
        errors: list[str] = []
        warnings: list[str] = []
        line_type = source["line_type"]
        if line_type not in LINE_TYPES:
            errors.append("INVALID_LINE_TYPE")
        account = accounts.get(source["account_code"])
        if not account:
            errors.append("ACCOUNT_NOT_FOUND")
        elif not account.active or not account.is_postable:
            errors.append("ACCOUNT_MUST_BE_ACTIVE_AND_POSTABLE")
        try:
            debit = money(source["debit_raw"])
            credit = money(source["credit_raw"])
        except ValueError as exc:
            errors.append(str(exc))
            debit, credit = Decimal("0"), Decimal("0")
        if debit < 0 or credit < 0:
            errors.append("NEGATIVE_AMOUNT_NOT_ALLOWED")
        if (debit > 0) == (credit > 0):
            errors.append("EXACTLY_ONE_OF_DEBIT_OR_CREDIT_IS_REQUIRED")
        party = parties.get(source["party_code"]) if source["party_code"] else None
        item = items.get(source["item_code"]) if source["item_code"] else None
        warehouse = warehouses.get(source["warehouse_code"]) if source["warehouse_code"] else None
        document_date = due_date = None
        qty = unit_cost = Decimal("0")
        try:
            document_date = parse_date(source["document_date_raw"], required=line_type in {"AR", "AP"})
            due_date = parse_date(source["due_date_raw"], required=line_type in {"AR", "AP"})
        except ValueError as exc:
            errors.append(str(exc))
        if document_date and due_date and due_date < document_date:
            errors.append("DUE_DATE_BEFORE_DOCUMENT_DATE")

        if line_type == "GL":
            if account and account.id in inventory_account_ids:
                errors.append("INVENTORY_CONTROL_ACCOUNT_REQUIRES_INVENTORY_LINE")
            if source["account_code"] in {"112010", "211010"}:
                errors.append("AR_AP_CONTROL_ACCOUNT_REQUIRES_SUBLEDGER_LINE")
        elif line_type in {"AR", "AP"}:
            if not party:
                errors.append("PARTY_NOT_FOUND")
            elif line_type == "AR" and party.party_type not in {"CUSTOMER", "BOTH"}:
                errors.append("AR_PARTY_MUST_BE_CUSTOMER")
            elif line_type == "AP" and party.party_type not in {"SUPPLIER", "BOTH"}:
                errors.append("AP_PARTY_MUST_BE_SUPPLIER")
            if line_type == "AR" and not (debit > 0 and credit == 0):
                errors.append("AR_OPEN_ITEM_MUST_BE_DEBIT")
            if line_type == "AP" and not (credit > 0 and debit == 0):
                errors.append("AP_OPEN_ITEM_MUST_BE_CREDIT")
            reference = source["reference_code"]
            if not reference:
                errors.append("DOCUMENT_REFERENCE_REQUIRED")
            else:
                key = (line_type, reference)
                if key in document_keys or reference in existing_documents:
                    errors.append("DUPLICATE_DOCUMENT_REFERENCE")
                document_keys.add(key)
        elif line_type == "INVENTORY":
            if not item:
                errors.append("ITEM_NOT_FOUND")
            if not warehouse:
                errors.append("WAREHOUSE_NOT_FOUND")
            if item and account and item.inventory_account_id != account.id:
                errors.append("ITEM_INVENTORY_ACCOUNT_MISMATCH")
            try:
                qty = quantity(source["quantity_raw"])
                unit_cost = quantity(source["unit_cost_raw"])
            except ValueError as exc:
                errors.append(str(exc))
            if qty <= 0 or unit_cost < 0:
                errors.append("POSITIVE_QUANTITY_AND_NONNEGATIVE_COST_REQUIRED")
            if credit > 0:
                errors.append("INVENTORY_OPENING_MUST_BE_DEBIT")
            if money(qty * unit_cost) != debit:
                errors.append("INVENTORY_VALUE_MUST_EQUAL_QUANTITY_TIMES_UNIT_COST")

        total_debit += debit
        total_credit += credit
        prepared.append({
            "line_number": source["line_number"], "line_type": line_type,
            "account": account, "party": party, "item": item, "warehouse": warehouse,
            "account_code": source["account_code"], "party_code": source["party_code"],
            "item_code": source["item_code"], "warehouse_code": source["warehouse_code"],
            "reference_code": source["reference_code"], "document_date": document_date, "due_date": due_date,
            "quantity": qty, "unit_cost": unit_cost, "lot_number": source["lot_number"],
            "debit": debit, "credit": credit, "description": source["description"],
            "errors": list(dict.fromkeys(errors)), "warnings": list(dict.fromkeys(warnings)),
        })

    total_debit = money(total_debit)
    total_credit = money(total_credit)
    global_errors: list[str] = []
    if total_debit <= 0:
        global_errors.append("OPENING_BALANCE_TOTAL_MUST_BE_POSITIVE")
    if total_debit != total_credit:
        global_errors.append("OPENING_BALANCES_ARE_NOT_BALANCED")
    error_count = sum(len(row["errors"]) for row in prepared) + len(global_errors)
    public_rows = [
        {
            "line_number": row["line_number"], "line_type": row["line_type"],
            "account_code": row["account_code"], "party_code": row["party_code"],
            "item_code": row["item_code"], "warehouse_code": row["warehouse_code"],
            "debit": row["debit"], "credit": row["credit"],
            "errors": row["errors"], "warnings": row["warnings"],
        }
        for row in prepared
    ]
    return {
        "valid": error_count == 0,
        "summary": {
            "rows": len(prepared), "total_debit": total_debit, "total_credit": total_credit,
            "difference": money(total_debit - total_credit), "errors": error_count,
            "gl_lines": sum(1 for row in prepared if row["line_type"] == "GL"),
            "ar_lines": sum(1 for row in prepared if row["line_type"] == "AR"),
            "ap_lines": sum(1 for row in prepared if row["line_type"] == "AP"),
            "inventory_lines": sum(1 for row in prepared if row["line_type"] == "INVENTORY"),
        },
        "global_errors": global_errors,
        "rows": public_rows,
        "_prepared": prepared,
    }


def _serialize(batch: OpeningBalanceBatch, include_lines: bool = False) -> dict:
    result = {
        "id": batch.id, "company_id": batch.company_id, "opening_date": batch.opening_date,
        "version": batch.version, "source_system": batch.source_system,
        "source_filename": batch.source_filename, "status": batch.status,
        "total_debit": batch.total_debit, "total_credit": batch.total_credit,
        "validation_hash": batch.validation_hash, "journal_id": batch.journal_id,
        "created_by": batch.created_by, "submitted_by": batch.submitted_by,
        "approved_by": batch.approved_by, "created_at": batch.created_at,
        "submitted_at": batch.submitted_at, "posted_at": batch.posted_at,
        "line_count": len(batch.lines),
    }
    if include_lines:
        result["lines"] = [
            {
                "id": line.id, "line_number": line.line_number, "line_type": line.line_type,
                "account_code": line.account.code, "account_name_ar": line.account.name_ar,
                "party_code": line.party.code if line.party else None,
                "item_code": line.item.code if line.item else None,
                "warehouse_code": line.warehouse.code if line.warehouse else None,
                "reference_code": line.reference_code, "document_date": line.document_date,
                "due_date": line.due_date, "quantity": line.quantity, "unit_cost": line.unit_cost,
                "debit": line.debit, "credit": line.credit, "description": line.description,
            }
            for line in batch.lines
        ]
    return result


def _get_batch(db: Session, batch_id: int) -> OpeningBalanceBatch:
    batch = db.scalar(
        select(OpeningBalanceBatch)
        .where(OpeningBalanceBatch.id == batch_id)
        .options(
            selectinload(OpeningBalanceBatch.lines).selectinload(OpeningBalanceLine.account),
            selectinload(OpeningBalanceBatch.lines).selectinload(OpeningBalanceLine.party),
            selectinload(OpeningBalanceBatch.lines).selectinload(OpeningBalanceLine.item),
            selectinload(OpeningBalanceBatch.lines).selectinload(OpeningBalanceLine.warehouse),
        )
    )
    if not batch:
        raise HTTPException(404, "Opening-balance batch not found")
    return batch


@router.get("/template.xlsx")
def template(company_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "finance.opening.read")
    accounts = db.scalars(select(Account).where(Account.company_id == company_id, Account.active.is_(True), Account.is_postable.is_(True)).order_by(Account.code)).all()
    parties = db.scalars(select(Party).where(Party.company_id == company_id, Party.active.is_(True)).order_by(Party.code)).all()
    items = db.scalars(select(Item).where(Item.company_id == company_id, Item.active.is_(True)).order_by(Item.code)).all()
    warehouses = db.scalars(select(Warehouse).where(Warehouse.company_id == company_id, Warehouse.active.is_(True)).order_by(Warehouse.code)).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Opening_Balances"
    sheet.append(list(HEADERS))
    sheet.freeze_panes = "A2"
    sheet.sheet_view.rightToLeft = True
    sheet.auto_filter.ref = "A1:N1"
    instructions = workbook.create_sheet("Instructions_AR")
    instructions.sheet_view.rightToLeft = True
    instructions.append(["تعليمات الاستيراد"])
    instructions.append(["GL", "رصيد حساب عادي. لا تستخدم GL لحسابات العملاء/الموردين/المخزون."])
    instructions.append(["AR", "فاتورة عميل مفتوحة: يلزم party_code وreference_code وتاريخ المستند والاستحقاق."])
    instructions.append(["AP", "فاتورة مورد مفتوحة: يلزم party_code وreference_code وتاريخ المستند والاستحقاق."])
    instructions.append(["INVENTORY", "رصيد صنف: يلزم item_code وwarehouse_code والكمية والتكلفة، ويجب أن يساوي المدين الكمية × التكلفة."])
    instructions.append(["قاعدة", "يجب أن يتساوى إجمالي المدين والدائن قبل السماح بالترحيل."])
    for title, headers, values in (
        ("Accounts", ["account_code", "name_ar", "account_type"], [(_excel_safe(r.code), _excel_safe(r.name_ar), r.account_type) for r in accounts]),
        ("Parties", ["party_code", "name_ar", "party_type"], [(_excel_safe(r.code), _excel_safe(r.name_ar), r.party_type) for r in parties]),
        ("Items", ["item_code", "name_ar", "inventory_account"], [(_excel_safe(r.code), _excel_safe(r.name_ar), _excel_safe(r.inventory_account.code)) for r in items]),
        ("Warehouses", ["warehouse_code", "name_ar"], [(_excel_safe(r.code), _excel_safe(r.name_ar)) for r in warehouses]),
    ):
        lookup = workbook.create_sheet(title)
        lookup.append(headers)
        for value in values:
            lookup.append(value)
        lookup.freeze_panes = "A2"
        lookup.sheet_view.rightToLeft = True
    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="17365D")
            cell.alignment = Alignment(horizontal="center")
        for column in worksheet.columns:
            letter = column[0].column_letter
            worksheet.column_dimensions[letter].width = min(42, max(14, max(len(str(c.value or "")) for c in column) + 2))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="CORVAX_Opening_Balances_{company_id}.xlsx"'},
    )


@router.post("/validate")
def validate_file(
    company_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_permission(db, user, company_id, "finance.opening.manage")
    result = _validate(db, company_id, _parse_workbook(file))
    result.pop("_prepared", None)
    return result


@router.post("/import", status_code=201)
def import_file(
    company_id: int,
    opening_date: date,
    source_system: str,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_permission(db, user, company_id, "finance.opening.manage")
    ensure_open_period(db, company_id, opening_date)
    parsed = _parse_workbook(file)
    result = _validate(db, company_id, parsed)
    if not result["valid"]:
        result.pop("_prepared", None)
        raise HTTPException(422, result)
    posted = db.scalar(select(OpeningBalanceBatch.id).where(
        OpeningBalanceBatch.company_id == company_id,
        OpeningBalanceBatch.status == "POSTED",
    ))
    if posted:
        raise HTTPException(409, "A posted opening-balance batch already exists for this company")
    version = int(db.scalar(select(func.coalesce(func.max(OpeningBalanceBatch.version), 0)).where(
        OpeningBalanceBatch.company_id == company_id,
        OpeningBalanceBatch.opening_date == opening_date,
    )) or 0) + 1
    prepared: list[dict] = result["_prepared"]
    hash_rows = [
        {
            "line_number": row["line_number"], "line_type": row["line_type"],
            "account_code": row["account_code"], "party_code": row["party_code"],
            "item_code": row["item_code"], "warehouse_code": row["warehouse_code"],
            "reference_code": row["reference_code"], "document_date": row["document_date"],
            "due_date": row["due_date"], "quantity": row["quantity"], "unit_cost": row["unit_cost"],
            "lot_number": row["lot_number"], "debit": row["debit"], "credit": row["credit"],
            "description": row["description"],
        }
        for row in prepared
    ]
    batch = OpeningBalanceBatch(
        company_id=company_id, opening_date=opening_date, version=version,
        source_system=source_system.strip() or "LEGACY_SYSTEM",
        source_filename=(file.filename or "opening_balances.xlsx")[:255],
        status="DRAFT", total_debit=result["summary"]["total_debit"],
        total_credit=result["summary"]["total_credit"],
        validation_hash=_hash_rows(hash_rows), created_by=user.id,
    )
    for row in prepared:
        batch.lines.append(OpeningBalanceLine(
            line_number=row["line_number"], line_type=row["line_type"], account_id=row["account"].id,
            party_id=row["party"].id if row["party"] else None,
            item_id=row["item"].id if row["item"] else None,
            warehouse_id=row["warehouse"].id if row["warehouse"] else None,
            reference_code=row["reference_code"], document_date=row["document_date"], due_date=row["due_date"],
            quantity=row["quantity"] or None, unit_cost=row["unit_cost"] or None, lot_number=row["lot_number"],
            debit=row["debit"], credit=row["credit"], description=row["description"],
        ))
    db.add(batch)
    db.flush()
    write_audit(
        db, action="OPENING_BALANCES_IMPORTED", entity_type="OPENING_BALANCE_BATCH", entity_id=batch.id,
        user_id=user.id, company_id=company_id,
        after={"date": str(opening_date), "version": version, "lines": len(batch.lines), "hash": batch.validation_hash},
    )
    db.commit()
    return _serialize(_get_batch(db, batch.id), include_lines=True)


@router.get("")
def list_batches(company_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "finance.opening.read")
    rows = db.scalars(
        select(OpeningBalanceBatch)
        .where(OpeningBalanceBatch.company_id == company_id)
        .options(selectinload(OpeningBalanceBatch.lines))
        .order_by(OpeningBalanceBatch.opening_date.desc(), OpeningBalanceBatch.version.desc())
    ).all()
    return [_serialize(row) for row in rows]


@router.get("/{batch_id}")
def get_batch(batch_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    batch = _get_batch(db, batch_id)
    ensure_permission(db, user, batch.company_id, "finance.opening.read")
    return _serialize(batch, include_lines=True)


@router.post("/{batch_id}/submit")
def submit_batch(batch_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    batch = _get_batch(db, batch_id)
    ensure_permission(db, user, batch.company_id, "finance.opening.manage")
    if batch.status != "DRAFT":
        raise HTTPException(409, "Only a draft opening-balance batch can be submitted")
    if _stored_hash(batch) != batch.validation_hash:
        raise HTTPException(409, "Opening-balance integrity hash mismatch")
    batch.status = "PENDING_APPROVAL"
    batch.submitted_by = user.id
    batch.submitted_at = utc_now()
    write_audit(
        db, action="OPENING_BALANCES_SUBMITTED", entity_type="OPENING_BALANCE_BATCH", entity_id=batch.id,
        user_id=user.id, company_id=batch.company_id, before={"status": "DRAFT"}, after={"status": batch.status},
    )
    db.commit()
    return _serialize(batch)


@router.post("/{batch_id}/approve-post")
def approve_post_batch(batch_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    batch = _get_batch(db, batch_id)
    ensure_permission(db, user, batch.company_id, "finance.opening.approve")
    if batch.status != "PENDING_APPROVAL":
        raise HTTPException(409, "Opening-balance batch must be PENDING_APPROVAL")
    if user.id in {batch.created_by, batch.submitted_by}:
        raise HTTPException(409, "The preparer/submitter cannot approve opening balances")
    if _stored_hash(batch) != batch.validation_hash:
        raise HTTPException(409, "Opening-balance integrity hash mismatch")
    if db.scalar(select(OpeningBalanceBatch.id).where(
        OpeningBalanceBatch.company_id == batch.company_id,
        OpeningBalanceBatch.status == "POSTED",
        OpeningBalanceBatch.id != batch.id,
    )):
        raise HTTPException(409, "A posted opening-balance batch already exists")
    ensure_open_period(db, batch.company_id, batch.opening_date)
    journal = create_posted_journal(
        db, company_id=batch.company_id, user_id=user.id, posting_date=batch.opening_date,
        reference=f"OPENING-{batch.company_id}-{batch.opening_date}-V{batch.version}",
        description=f"Controlled opening balances from {batch.source_system}",
        lines=[
            {
                "account_id": line.account_id, "debit": line.debit, "credit": line.credit,
                "description": line.description or f"Opening {line.line_type} line {line.line_number}",
            }
            for line in batch.lines
        ],
    )
    journal.entry_origin = "OPENING_BALANCE"
    for line in batch.lines:
        if line.line_type in {"AR", "AP"}:
            db.add(FinancialOpenItem(
                company_id=batch.company_id, ledger_type=line.line_type, party_id=line.party_id,
                source_type="OPENING_BALANCE", source_id=line.id,
                document_number=line.reference_code, document_date=line.document_date,
                due_date=line.due_date,
                original_amount=money(line.debit if line.line_type == "AR" else line.credit),
                status="OPEN", journal_id=journal.id,
                notes=line.description or f"Imported from {batch.source_system}", created_by=user.id,
            ))
        elif line.line_type == "INVENTORY":
            prior = db.scalar(select(StockMovement.id).where(
                StockMovement.company_id == batch.company_id,
                StockMovement.item_id == line.item_id,
                StockMovement.warehouse_id == line.warehouse_id,
                StockMovement.movement_date <= batch.opening_date,
            ))
            if prior:
                raise HTTPException(409, f"Prior stock movement prevents opening line {line.line_number}")
            db.add(StockMovement(
                company_id=batch.company_id, warehouse_id=line.warehouse_id, item_id=line.item_id,
                movement_date=batch.opening_date, movement_type="OPENING",
                quantity=line.quantity, unit_cost=line.unit_cost, total_cost=line.debit,
                lot_number=line.lot_number, reference_type="OPENING_BALANCE",
                reference_id=batch.id, journal_id=journal.id, created_by=user.id,
            ))
    batch.status = "POSTED"
    batch.approved_by = user.id
    batch.posted_at = utc_now()
    batch.journal_id = journal.id
    write_audit(
        db, action="OPENING_BALANCES_APPROVED_POSTED", entity_type="OPENING_BALANCE_BATCH", entity_id=batch.id,
        user_id=user.id, company_id=batch.company_id,
        after={"journal": journal.number, "total": str(batch.total_debit), "hash": batch.validation_hash},
    )
    db.commit()
    return {**_serialize(batch), "journal_number": journal.number}
