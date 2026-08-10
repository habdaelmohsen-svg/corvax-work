"""R9 platform assurance API.

The router is intentionally not wired into ``main.py`` in this change set.  See
``R9_PLATFORM_INTEGRATION.md`` for the explicit, reviewable integration steps.
"""
from __future__ import annotations

import base64
import hashlib
import io
import json
import re
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import func, or_, select, text
from sqlalchemy.orm import Session, selectinload

from app.core.config import settings
from app.core.time import utc_now
from app.db import engine, get_db
from app.dependencies import ensure_permission, get_current_user
from app.models.audit_banking import AuditLog
from app.models.core import Role, User, UserCompanyRole
from app.models.operations_compliance import BackupRecord
from app.models.qms_food_access import SoDConflict
from app.models.r9_platform import (
    R9ImportBatch, R9ImportRow, R9PlatformAlert, R9RestoreDrill,
    R9ZatcaReadiness, R9ZatcaSandboxSubmission,
)
from app.services.audit import write_audit

router = APIRouter(prefix="/r9-platform", tags=["R9 platform assurance"])

MAX_XLSX_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
TARGET_RULES: dict[str, dict[str, Any]] = {
    "SUPPLIERS": {"required": {"code", "name", "vat_number"}, "decimal": set()},
    "CUSTOMERS": {"required": {"code", "name"}, "decimal": {"credit_limit"}},
    "ITEMS": {"required": {"code", "name", "unit"}, "decimal": {"standard_cost", "selling_price"}},
    "OPENING_BALANCES": {"required": {"account_code", "debit", "credit"}, "decimal": {"debit", "credit"}},
}


def _dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str, sort_keys=True, separators=(",", ":"))


def _load(value: str | None, fallback: Any) -> Any:
    try:
        return json.loads(value) if value else fallback
    except json.JSONDecodeError:
        return fallback


def _migration_head(db: Session) -> str:
    try:
        return str(db.execute(text("SELECT version_num FROM alembic_version")).scalar_one())
    except Exception:
        db.rollback()
        return "UNAVAILABLE"


def _pool_metrics() -> dict[str, int | str]:
    pool = engine.pool
    result: dict[str, int | str] = {"class": pool.__class__.__name__}
    for name in ("size", "checkedin", "checkedout", "overflow"):
        method = getattr(pool, name, None)
        if callable(method):
            try:
                result[name] = int(method())
            except Exception:
                result[name] = "UNAVAILABLE"
    return result


@router.get("/health")
def platform_health(company_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "platform.view")
    now = utc_now()
    latest_backup = db.scalar(select(BackupRecord).where(BackupRecord.company_id == company_id).order_by(BackupRecord.created_at.desc()).limit(1))
    latest_drill = db.scalar(select(R9RestoreDrill).where(R9RestoreDrill.company_id == company_id).order_by(R9RestoreDrill.performed_at.desc()).limit(1))
    errors_24h = db.scalar(select(func.count(AuditLog.id)).where(
        AuditLog.company_id == company_id,
        AuditLog.created_at >= now - timedelta(hours=24),
        or_(func.upper(AuditLog.action).like("%FAIL%"), func.upper(AuditLog.action).like("%ERROR%"),
            func.upper(AuditLog.action).like("%EXCEPTION%")),
    )) or 0
    open_alerts = db.scalar(select(func.count(R9PlatformAlert.id)).where(
        R9PlatformAlert.company_id == company_id, R9PlatformAlert.status.in_(["OPEN", "ASSIGNED"])
    )) or 0
    return {
        "service": {"status": "UP", "version": settings.app_version, "environment": settings.environment},
        "database": {"status": "UP", "driver": engine.dialect.name, "migration_head": _migration_head(db), "pool": _pool_metrics()},
        "errors": {"failed_audit_events_24h": errors_24h, "open_control_alerts": open_alerts},
        "backup": {
            "last_status": latest_backup.status if latest_backup else "NEVER",
            "last_created_at": latest_backup.created_at if latest_backup else None,
            "last_verified_at": latest_backup.verified_at if latest_backup else None,
            "storage_detail_exposed": False,
        },
        "restore_drill": {
            "last_status": latest_drill.status if latest_drill else "NEVER",
            "last_performed_at": latest_drill.performed_at if latest_drill else None,
            "environment": latest_drill.environment if latest_drill else None,
        },
        "generated_at": now,
        "secrets_exposed": False,
    }


class RestoreDrillIn(BaseModel):
    backup_id: int
    status: str
    integrity_check: str = Field(min_length=2, max_length=40)
    notes: str = Field(min_length=10, max_length=2000)


@router.post("/restore-drills", status_code=201)
def record_restore_drill(data: RestoreDrillIn, company_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "platform.manage")
    backup = db.scalar(select(BackupRecord).where(BackupRecord.id == data.backup_id, BackupRecord.company_id == company_id))
    if not backup or backup.status not in {"VERIFIED", "COMPLETED"}:
        raise HTTPException(422, "A company-scoped completed or verified backup is required")
    status = data.status.upper()
    if status not in {"PASSED", "FAILED"}:
        raise HTTPException(422, "Status must be PASSED or FAILED")
    if status == "PASSED" and backup.status != "VERIFIED":
        raise HTTPException(422, "A passed restore drill requires a checksum-verified backup")
    row = R9RestoreDrill(company_id=company_id, backup_id=backup.id, status=status,
                         integrity_check=data.integrity_check.upper(), notes=data.notes,
                         environment="ISOLATED_TEST", performed_by=user.id)
    db.add(row); db.flush()
    write_audit(db, action="R9_RESTORE_DRILL_RECORDED", entity_type="RESTORE_DRILL", entity_id=row.id,
                user_id=user.id, company_id=company_id, after={"backup_id": backup.id, "status": status, "environment": row.environment})
    db.commit()
    return {"id": row.id, "status": row.status, "environment": row.environment, "production_restore": False}


def _alert_fingerprint(company_id: int, code: str, entity: str = "") -> str:
    return hashlib.sha256(f"{company_id}:{code}:{entity}".encode()).hexdigest()


def _upsert_alert(db: Session, *, company_id: int, code: str, category: str, severity: str,
                  title_ar: str, title_en: str, details: dict[str, Any], entity_type: str | None = None,
                  entity_id: str | int | None = None) -> bool:
    fingerprint = _alert_fingerprint(company_id, code, str(entity_id or ""))
    row = db.scalar(select(R9PlatformAlert).where(
        R9PlatformAlert.company_id == company_id, R9PlatformAlert.fingerprint == fingerprint,
    ))
    if row:
        row.details_json = _dump(details); row.updated_at = utc_now()
        if row.status == "RESOLVED":
            row.status = "OPEN"; row.assigned_to = None; row.resolved_by = None
            row.resolution_notes = None; row.resolved_at = None; row.detected_at = utc_now()
            return True
        return False
    row = R9PlatformAlert(company_id=company_id, fingerprint=fingerprint, category=category,
                          severity=severity, title_ar=title_ar, title_en=title_en,
                          details_json=_dump(details), source_entity_type=entity_type,
                          source_entity_id=str(entity_id) if entity_id is not None else None)
    db.add(row)
    return True


@router.post("/controls/scan")
def scan_controls(company_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "platform.manage")
    created = 0
    conflicts = db.scalars(select(SoDConflict).where(
        SoDConflict.company_id == company_id, SoDConflict.status.in_(["OPEN", "MITIGATED"])
    )).all()
    for conflict in conflicts:
        created += int(_upsert_alert(db, company_id=company_id, code="SOD_CONFLICT", category="SOD",
            severity="CRITICAL" if conflict.status == "OPEN" else "HIGH",
            title_ar="تعارض قائم في فصل المهام", title_en="Active segregation-of-duties conflict",
            details={"status": conflict.status, "user_id": conflict.user_id, "rule_id": conflict.rule_id},
            entity_type="SOD_CONFLICT", entity_id=conflict.id))

    memberships = db.scalars(select(UserCompanyRole).options(
        selectinload(UserCompanyRole.user), selectinload(UserCompanyRole.role)
    ).where(UserCompanyRole.company_id == company_id)).all()
    for membership in memberships:
        if membership.role.code in settings.sensitive_roles and not membership.user.mfa_enabled:
            created += int(_upsert_alert(db, company_id=company_id, code="SENSITIVE_MFA_DISABLED", category="SECURITY",
                severity="CRITICAL", title_ar="المصادقة المتعددة غير مفعلة لدور حساس",
                title_en="MFA is disabled for a sensitive role",
                details={"user_id": membership.user_id, "role": membership.role.code},
                entity_type="USER", entity_id=membership.user_id))

    latest_backup = db.scalar(select(BackupRecord).where(BackupRecord.company_id == company_id).order_by(BackupRecord.created_at.desc()).limit(1))
    if not latest_backup or latest_backup.created_at < utc_now() - timedelta(days=7):
        created += int(_upsert_alert(db, company_id=company_id, code="BACKUP_OVERDUE", category="RESILIENCE", severity="HIGH",
            title_ar="لا توجد نسخة احتياطية حديثة", title_en="A recent backup is missing",
            details={"maximum_age_days": 7, "last_backup_at": latest_backup.created_at if latest_backup else None}))
    latest_drill = db.scalar(select(R9RestoreDrill).where(R9RestoreDrill.company_id == company_id).order_by(R9RestoreDrill.performed_at.desc()).limit(1))
    if not latest_drill or latest_drill.performed_at < utc_now() - timedelta(days=30) or latest_drill.status != "PASSED":
        created += int(_upsert_alert(db, company_id=company_id, code="RESTORE_DRILL_DUE", category="RESILIENCE", severity="HIGH",
            title_ar="اختبار استعادة ناجح مطلوب", title_en="A successful restore drill is due",
            details={"maximum_age_days": 30, "last_status": latest_drill.status if latest_drill else "NEVER"}))
    write_audit(db, action="R9_CONTROL_SCAN_EXECUTED", entity_type="COMPANY", entity_id=company_id,
                user_id=user.id, company_id=company_id,
                after={"created_alerts": created, "sod_conflicts_scanned": len(conflicts), "memberships_scanned": len(memberships)})
    db.commit()
    total = db.scalar(select(func.count(R9PlatformAlert.id)).where(
        R9PlatformAlert.company_id == company_id, R9PlatformAlert.status.in_(["OPEN", "ASSIGNED"]))) or 0
    return {"company_id": company_id, "created_alerts": created, "open_alerts": total}


@router.get("/alerts")
def list_alerts(company_id: int, status: str | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "platform.view")
    query = select(R9PlatformAlert).where(R9PlatformAlert.company_id == company_id)
    if status:
        query = query.where(R9PlatformAlert.status == status.upper())
    rows = db.scalars(query.order_by(R9PlatformAlert.detected_at.desc())).all()
    return [{"id": r.id, "category": r.category, "severity": r.severity, "title_ar": r.title_ar,
             "title_en": r.title_en, "status": r.status, "details": _load(r.details_json, {}),
             "assigned_to": r.assigned_to, "detected_at": r.detected_at} for r in rows]


class AlertDecisionIn(BaseModel):
    assigned_to: int | None = None
    resolution_notes: str | None = Field(default=None, max_length=2000)


@router.post("/alerts/{alert_id}/assign")
def assign_alert(alert_id: int, data: AlertDecisionIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.get(R9PlatformAlert, alert_id)
    if not row: raise HTTPException(404, "Alert not found")
    ensure_permission(db, user, row.company_id, "platform.manage")
    assignee = data.assigned_to or user.id
    membership = db.scalar(select(UserCompanyRole.id).where(UserCompanyRole.company_id == row.company_id, UserCompanyRole.user_id == assignee))
    if not membership: raise HTTPException(422, "Assignee has no access to this company")
    row.assigned_to = assignee; row.status = "ASSIGNED"; row.updated_at = utc_now()
    write_audit(db, action="R9_ALERT_ASSIGNED", entity_type="PLATFORM_ALERT", entity_id=row.id,
                user_id=user.id, company_id=row.company_id, after={"assigned_to": assignee})
    db.commit(); return {"id": row.id, "status": row.status, "assigned_to": assignee}


@router.post("/alerts/{alert_id}/resolve")
def resolve_alert(alert_id: int, data: AlertDecisionIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.get(R9PlatformAlert, alert_id)
    if not row: raise HTTPException(404, "Alert not found")
    ensure_permission(db, user, row.company_id, "platform.manage")
    if not data.resolution_notes or len(data.resolution_notes.strip()) < 10:
        raise HTTPException(422, "Resolution evidence must contain at least 10 characters")
    before = {"status": row.status, "assigned_to": row.assigned_to}
    row.status = "RESOLVED"; row.resolved_by = user.id; row.resolution_notes = data.resolution_notes
    row.resolved_at = utc_now(); row.updated_at = row.resolved_at
    write_audit(db, action="R9_ALERT_RESOLVED", entity_type="PLATFORM_ALERT", entity_id=row.id,
                user_id=user.id, company_id=row.company_id, before=before,
                after={"status": row.status, "resolution_notes": data.resolution_notes})
    db.commit(); return {"id": row.id, "status": row.status, "resolved_at": row.resolved_at}


def _clean_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9_]+", "_", str(value or "").strip().lower()).strip("_")


def parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > 2_000 or sum(item.file_size for item in entries) > 50 * 1024 * 1024:
                raise HTTPException(413, "Expanded workbook exceeds the safety limit")
            names = {item.filename.lower() for item in entries}
            if any(name.endswith("vbaproject.bin") for name in names):
                raise HTTPException(422, "Macro-enabled content is not accepted")
    except HTTPException:
        raise
    except zipfile.BadZipFile as exc:
        raise HTTPException(422, "Invalid XLSX container") from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(422, "Invalid or unsupported XLSX workbook") from exc
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    raw_headers = next(iterator, None)
    if not raw_headers: raise HTTPException(422, "Workbook is empty")
    headers = [_clean_header(value) for value in raw_headers]
    if any(not h for h in headers) or len(headers) != len(set(headers)):
        raise HTTPException(422, "Headers must be non-empty and unique")
    rows: list[dict[str, Any]] = []
    for number, values in enumerate(iterator, start=2):
        if number > MAX_IMPORT_ROWS + 1: raise HTTPException(413, f"Workbook exceeds {MAX_IMPORT_ROWS} data rows")
        if all(value is None or str(value).strip() == "" for value in values): continue
        rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    workbook.close()
    return headers, rows


@router.post("/imports/stage", status_code=201)
async def stage_import(company_id: int, target_type: str, file: UploadFile = File(...),
                       user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "import.stage")
    target = target_type.upper()
    if target not in TARGET_RULES: raise HTTPException(422, f"Unsupported target type: {target}")
    filename = (file.filename or "").rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    if not filename.lower().endswith(".xlsx"): raise HTTPException(422, "Only non-macro XLSX workbooks are accepted")
    content = await file.read(MAX_XLSX_BYTES + 1)
    if len(content) > MAX_XLSX_BYTES: raise HTTPException(413, "Workbook exceeds 5 MiB")
    headers, rows = parse_xlsx(content)
    missing = sorted(TARGET_RULES[target]["required"] - set(headers))
    if missing: raise HTTPException(422, {"message": "Required columns are missing", "columns": missing})
    digest = hashlib.sha256(content).hexdigest()
    if db.scalar(select(R9ImportBatch.id).where(R9ImportBatch.company_id == company_id,
        R9ImportBatch.target_type == target, R9ImportBatch.file_sha256 == digest)):
        raise HTTPException(409, "This workbook was already staged for the same target")
    batch = R9ImportBatch(company_id=company_id, target_type=target, original_filename=filename,
                          file_sha256=digest, total_rows=len(rows), created_by=user.id)
    db.add(batch); db.flush()
    for number, payload in enumerate(rows, start=2):
        db.add(R9ImportRow(batch_id=batch.id, row_number=number, payload_json=_dump(payload)))
    write_audit(db, action="R9_IMPORT_STAGED", entity_type="IMPORT_BATCH", entity_id=batch.id,
                user_id=user.id, company_id=company_id,
                after={"target_type": target, "file_sha256": digest, "rows": len(rows), "filename": filename})
    db.commit()
    return {"id": batch.id, "status": batch.status, "target_type": target, "total_rows": batch.total_rows,
            "message": "Staged only; no master or financial records were posted"}


def _validate_row(target: str, payload: dict[str, Any], seen_codes: set[str]) -> tuple[dict[str, Any], list[str]]:
    rules = TARGET_RULES[target]; normalized: dict[str, Any] = {}; errors: list[str] = []
    for key, value in payload.items():
        normalized[key] = value.strip() if isinstance(value, str) else value
    for key in rules["required"]:
        if normalized.get(key) is None or str(normalized[key]).strip() == "": errors.append(f"{key}: required")
    code_key = "account_code" if target == "OPENING_BALANCES" else "code"
    code = str(normalized.get(code_key) or "").upper()
    normalized[code_key] = code
    if code in seen_codes: errors.append(f"{code_key}: duplicate within workbook")
    if code: seen_codes.add(code)
    for key in rules["decimal"]:
        raw = normalized.get(key)
        if raw in (None, ""): normalized[key] = "0.00"; continue
        try: normalized[key] = str(Decimal(str(raw)).quantize(Decimal("0.01")))
        except InvalidOperation: errors.append(f"{key}: invalid decimal")
        else:
            if target in {"CUSTOMERS", "ITEMS"} and Decimal(normalized[key]) < 0:
                errors.append(f"{key}: cannot be negative")
    if target == "SUPPLIERS":
        vat = re.sub(r"\D", "", str(normalized.get("vat_number") or ""))
        normalized["vat_number"] = vat
        if len(vat) != 15: errors.append("vat_number: must contain 15 digits")
    if target == "OPENING_BALANCES":
        try:
            if Decimal(normalized.get("debit", "0")) and Decimal(normalized.get("credit", "0")):
                errors.append("debit/credit: a row cannot contain both")
        except InvalidOperation: pass
    return normalized, errors


@router.post("/imports/{batch_id}/validate")
def validate_import(batch_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    batch = db.get(R9ImportBatch, batch_id)
    if not batch: raise HTTPException(404, "Import batch not found")
    ensure_permission(db, user, batch.company_id, "import.stage")
    if batch.status not in {"STAGED", "VALIDATION_FAILED"}: raise HTTPException(409, "Batch cannot be validated in its current status")
    rows = db.scalars(select(R9ImportRow).where(R9ImportRow.batch_id == batch.id).order_by(R9ImportRow.row_number)).all()
    seen: set[str] = set(); invalid = 0; error_counts: dict[str, int] = {}; batch_errors: list[str] = []
    for row in rows:
        normalized, errors = _validate_row(batch.target_type, _load(row.payload_json, {}), seen)
        row.normalized_json = _dump(normalized); row.errors_json = _dump(errors)
        row.validation_status = "INVALID" if errors else "VALID"
        if errors:
            invalid += 1
            for error in errors: error_counts[error.split(":", 1)[0]] = error_counts.get(error.split(":", 1)[0], 0) + 1
    if batch.target_type == "OPENING_BALANCES" and not invalid:
        debit = sum(Decimal(_load(row.normalized_json, {}).get("debit", "0")) for row in rows)
        credit = sum(Decimal(_load(row.normalized_json, {}).get("credit", "0")) for row in rows)
        if debit != credit:
            batch_errors.append(f"opening balance is not balanced: debit={debit}, credit={credit}")
    batch.valid_rows = len(rows) - invalid; batch.invalid_rows = invalid
    batch.status = "VALIDATION_FAILED" if invalid or batch_errors else "VALIDATED"
    batch.validated_by = user.id; batch.validated_at = utc_now()
    batch.validation_summary_json = _dump({"error_fields": error_counts, "batch_errors": batch_errors})
    write_audit(db, action="R9_IMPORT_VALIDATED", entity_type="IMPORT_BATCH", entity_id=batch.id,
                user_id=user.id, company_id=batch.company_id,
                after={"status": batch.status, "valid_rows": batch.valid_rows, "invalid_rows": invalid})
    db.commit()
    return {"id": batch.id, "status": batch.status, "valid_rows": batch.valid_rows,
            "invalid_rows": batch.invalid_rows, "summary": _load(batch.validation_summary_json, {})}


@router.post("/imports/{batch_id}/approve")
def approve_import(batch_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    batch = db.get(R9ImportBatch, batch_id)
    if not batch: raise HTTPException(404, "Import batch not found")
    ensure_permission(db, user, batch.company_id, "import.approve")
    if batch.created_by == user.id: raise HTTPException(409, "Maker-checker: batch creator cannot approve")
    if batch.status != "VALIDATED" or batch.invalid_rows:
        raise HTTPException(422, "Only a fully validated batch can be approved")
    batch.status = "APPROVED_STAGING_ONLY"; batch.approved_by = user.id; batch.approved_at = utc_now()
    write_audit(db, action="R9_IMPORT_STAGING_APPROVED", entity_type="IMPORT_BATCH", entity_id=batch.id,
                user_id=user.id, company_id=batch.company_id,
                after={"status": batch.status, "target_type": batch.target_type, "posted_to_master": False})
    db.commit()
    return {"id": batch.id, "status": batch.status, "posted_to_master": False,
            "next_step": "A target-specific posting adapter and separate approval are required"}


@router.get("/imports/{batch_id}")
def get_import(batch_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    batch = db.get(R9ImportBatch, batch_id)
    if not batch: raise HTTPException(404, "Import batch not found")
    ensure_permission(db, user, batch.company_id, "platform.view")
    rows = db.scalars(select(R9ImportRow).where(R9ImportRow.batch_id == batch.id).order_by(R9ImportRow.row_number).limit(200)).all()
    return {"id": batch.id, "target_type": batch.target_type, "status": batch.status,
            "total_rows": batch.total_rows, "valid_rows": batch.valid_rows, "invalid_rows": batch.invalid_rows,
            "preview_limited_to": 200, "posted_to_master": False,
            "rows": [{"row_number": r.row_number, "status": r.validation_status,
                      "normalized": _load(r.normalized_json, _load(r.payload_json, {})),
                      "errors": _load(r.errors_json, [])} for r in rows]}


class ZatcaReadinessIn(BaseModel):
    onboarding_status: str = "NOT_STARTED"
    seller_identity_ready: bool = False
    certificate_configured: bool = False
    signing_key_configured: bool = False
    sdk_validation_ready: bool = False
    notes: str | None = Field(default=None, max_length=2000)


@router.put("/zatca/readiness")
def update_zatca_readiness(company_id: int, data: ZatcaReadinessIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "zatca.manage")
    status = data.onboarding_status.upper()
    if status not in {"NOT_STARTED", "PREPARING", "SANDBOX_READY", "SANDBOX_TESTED"}:
        raise HTTPException(422, "Production onboarding states are not accepted by the R9 readiness register")
    row = db.scalar(select(R9ZatcaReadiness).where(R9ZatcaReadiness.company_id == company_id))
    before = None
    if not row:
        row = R9ZatcaReadiness(company_id=company_id, updated_by=user.id); db.add(row)
    else:
        before = {"onboarding_status": row.onboarding_status, "sdk_validation_ready": row.sdk_validation_ready}
    row.onboarding_status = status; row.environment = "SANDBOX"; row.production_connected = False
    row.seller_identity_ready = data.seller_identity_ready
    row.certificate_configured = data.certificate_configured
    row.signing_key_configured = data.signing_key_configured
    row.sdk_validation_ready = data.sdk_validation_ready
    row.notes = data.notes; row.updated_by = user.id; row.updated_at = utc_now()
    write_audit(db, action="R9_ZATCA_READINESS_UPDATED", entity_type="ZATCA_READINESS", entity_id=company_id,
                user_id=user.id, company_id=company_id, before=before,
                after={"onboarding_status": status, "environment": "SANDBOX", "production_connected": False})
    db.commit()
    return _zatca_readiness_view(row)


def _zatca_readiness_view(row: R9ZatcaReadiness | None) -> dict[str, Any]:
    if not row:
        return {"onboarding_status": "NOT_STARTED", "environment": "SANDBOX", "production_connected": False,
                "seller_identity_ready": False, "certificate_configured": False,
                "signing_key_configured": False, "sdk_validation_ready": False}
    return {"onboarding_status": row.onboarding_status, "environment": row.environment,
            "production_connected": False, "seller_identity_ready": row.seller_identity_ready,
            "certificate_configured": row.certificate_configured,
            "signing_key_configured": row.signing_key_configured, "sdk_validation_ready": row.sdk_validation_ready,
            "last_validation_at": row.last_validation_at, "updated_at": row.updated_at}


@router.get("/zatca/readiness")
def get_zatca_readiness(company_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "platform.view")
    return _zatca_readiness_view(db.scalar(select(R9ZatcaReadiness).where(R9ZatcaReadiness.company_id == company_id)))


class ZatcaDocumentIn(BaseModel):
    source_type: str = Field(min_length=2, max_length=40)
    source_id: str = Field(min_length=1, max_length=80)
    seller_name: str = Field(min_length=2, max_length=250)
    vat_number: str = Field(min_length=15, max_length=30)
    issue_datetime: datetime
    total_with_vat: Decimal = Field(ge=0)
    vat_total: Decimal = Field(ge=0)
    canonical_xml: str = Field(min_length=10, max_length=2_000_000)
    previous_invoice_hash: str | None = Field(default=None, max_length=64)

    @field_validator("issue_datetime")
    @classmethod
    def require_timezone(cls, value: datetime) -> datetime:
        if value.tzinfo is None or value.utcoffset() is None:
            raise ValueError("issue_datetime must include a timezone offset")
        return value


def _tlv(tags: list[tuple[int, str]]) -> str:
    output = bytearray()
    for tag, value in tags:
        encoded = value.encode("utf-8")
        if len(encoded) > 255: raise HTTPException(422, "QR metadata field exceeds TLV limit")
        output.extend((tag, len(encoded))); output.extend(encoded)
    return base64.b64encode(bytes(output)).decode()


@router.post("/zatca/documents", status_code=201)
def create_zatca_sandbox_document(company_id: int, data: ZatcaDocumentIn,
                                  user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_permission(db, user, company_id, "zatca.manage")
    readiness = db.scalar(select(R9ZatcaReadiness).where(R9ZatcaReadiness.company_id == company_id))
    errors: list[str] = []
    vat = re.sub(r"\D", "", data.vat_number)
    if len(vat) != 15: errors.append("VAT_NUMBER_MUST_HAVE_15_DIGITS")
    if data.vat_total > data.total_with_vat: errors.append("VAT_EXCEEDS_TOTAL")
    if not data.canonical_xml.lstrip().startswith("<"): errors.append("CANONICAL_XML_REQUIRED")
    if data.previous_invoice_hash and not re.fullmatch(r"[0-9a-fA-F]{64}", data.previous_invoice_hash):
        errors.append("PREVIOUS_HASH_INVALID")
    if not readiness or not readiness.sdk_validation_ready: errors.append("SDK_VALIDATION_NOT_CONFIRMED")
    invoice_uuid = str(uuid.uuid4())
    invoice_hash = hashlib.sha256(data.canonical_xml.encode("utf-8")).hexdigest()
    qr = _tlv([(1, data.seller_name), (2, vat), (3, data.issue_datetime.astimezone(timezone.utc).isoformat()),
               (4, str(data.total_with_vat)), (5, str(data.vat_total))])
    row = R9ZatcaSandboxSubmission(company_id=company_id, source_type=data.source_type.upper(),
        source_id=data.source_id, invoice_uuid=invoice_uuid, invoice_hash=invoice_hash,
        previous_invoice_hash=data.previous_invoice_hash, qr_metadata_base64=qr,
        validation_status="INVALID" if errors else "INTERNALLY_VALIDATED",
        validation_errors_json=_dump(errors), submission_status="NOT_SUBMITTED", created_by=user.id)
    db.add(row); db.flush()
    if readiness:
        readiness.last_validation_at = utc_now()
    write_audit(db, action="R9_ZATCA_SANDBOX_DOCUMENT_CREATED", entity_type="ZATCA_SANDBOX_DOCUMENT", entity_id=row.id,
                user_id=user.id, company_id=company_id,
                after={"uuid": invoice_uuid, "hash": invoice_hash, "validation_status": row.validation_status,
                       "environment": "SANDBOX", "submitted": False})
    db.commit()
    return {"id": row.id, "uuid": row.invoice_uuid, "invoice_hash": row.invoice_hash,
            "qr_metadata_base64": row.qr_metadata_base64, "validation_status": row.validation_status,
            "validation_errors": errors, "submission_status": "NOT_SUBMITTED",
            "environment": "SANDBOX", "production_connected": False,
            "disclaimer": "Internal readiness evidence only; not a ZATCA acceptance or clearance"}


class SandboxEvidenceIn(BaseModel):
    correlation_id: str = Field(min_length=3, max_length=100)
    result: str


@router.post("/zatca/documents/{document_id}/sandbox-evidence")
def record_sandbox_evidence(document_id: int, data: SandboxEvidenceIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.get(R9ZatcaSandboxSubmission, document_id)
    if not row: raise HTTPException(404, "Sandbox document not found")
    ensure_permission(db, user, row.company_id, "zatca.manage")
    result = data.result.upper()
    if result not in {"SANDBOX_ACCEPTED", "SANDBOX_REJECTED"}: raise HTTPException(422, "Only sandbox evidence states are accepted")
    if row.validation_status != "INTERNALLY_VALIDATED": raise HTTPException(409, "Invalid document cannot receive accepted sandbox evidence")
    row.submission_status = result; row.sandbox_correlation_id = data.correlation_id
    write_audit(db, action="R9_ZATCA_SANDBOX_EVIDENCE_RECORDED", entity_type="ZATCA_SANDBOX_DOCUMENT", entity_id=row.id,
                user_id=user.id, company_id=row.company_id,
                after={"submission_status": result, "correlation_id": data.correlation_id,
                       "evidence_source": "MANUAL_SANDBOX_EVIDENCE", "production": False})
    db.commit()
    return {"id": row.id, "submission_status": result, "environment": "SANDBOX",
            "production_connected": False, "evidence_source": "MANUAL_SANDBOX_EVIDENCE"}
