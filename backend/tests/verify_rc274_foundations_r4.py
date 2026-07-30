"""CORVAX RC27.4 R4 verification for the six owner findings."""
from __future__ import annotations

import os
import subprocess
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook


BACKEND_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = BACKEND_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))
DB_PATH = Path("/tmp/verify_rc274_foundations_r4.db")
DB_PATH.unlink(missing_ok=True)
os.environ.update({
    "DATABASE_URL": f"sqlite:///{DB_PATH}",
    "SECRET_KEY": "verification-secret-key-corvax-rc274-foundations-r4",
    "SEED_DEMO_DATA": "true",
    "AUTO_CREATE_SCHEMA": "true",
    "TRUSTED_HOSTS": "testserver,localhost,127.0.0.1",
    "APP_VERSION": "1.0.0-agreement-completion-rc27.4-r4",
    "ENABLE_RATE_LIMIT_TESTING": "true",
    "PAYROLL_STRICT_WORKFLOW": "true",
})
subprocess.run([sys.executable, "-m", "alembic", "upgrade", "head"], cwd=BACKEND_DIR, check=True)

from fastapi.testclient import TestClient  # noqa: E402
from sqlalchemy import select, text  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    Account, Employee, Item, JournalEntry, OpeningBalanceBatch, User,
)


def ok(response, status=200):
    assert response.status_code == status, response.text
    return response.json()


def workbook_bytes(headers: list[str], rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def main() -> None:
    with TestClient(app) as client:
        with SessionLocal() as db:
            admin_user = db.scalar(select(User).where(User.email == "admin@corvaxplatform.com"))
            assert admin_user
            admin_user.require_password_change = False
            db.execute(text("update fiscal_periods set status='OPEN'"))
            db.commit()

        login = ok(client.post("/api/v1/auth/login", json={
            "email": "admin@corvaxplatform.com", "password": "Corvax@123",
        }))
        admin = {"Authorization": f"Bearer {login['access_token']}"}
        ok(client.post("/api/v1/admin/users", headers=admin, json={
            "name_ar": "معتمد الأرصدة", "name_en": "Opening Approver",
            "email": "r4.approver@corvaxplatform.com", "password": "R4Approver@123",
            "require_password_change": False,
            "memberships": [{"company_id": 1, "role_code": "SUPER_ADMIN"}],
        }), 201)
        approver_login = ok(client.post("/api/v1/auth/login", json={
            "email": "r4.approver@corvaxplatform.com", "password": "R4Approver@123",
        }))
        approver = {"Authorization": f"Bearer {approver_login['access_token']}"}

        # 1) COA exports a real XLSX and imports only after a clean validation.
        exported = client.get("/api/v1/chart-of-accounts/export.xlsx?company_id=1", headers=admin)
        assert exported.status_code == 200
        exported_book = load_workbook(BytesIO(exported.content), read_only=True)
        assert exported_book.active.title == "Chart_of_Accounts"
        assert exported_book.active.max_row > 2
        with SessionLocal() as db:
            parent = db.scalar(select(Account).where(
                Account.company_id == 1, Account.parent_id.is_(None), Account.account_type == "ASSET",
            ).order_by(Account.code))
            assert parent
            import_code = f"{parent.code.rstrip('0') or '1'}990001"
            assert not db.scalar(select(Account.id).where(Account.company_id == 1, Account.code == import_code))
            parent_code, account_type, statement_group = parent.code, parent.account_type, parent.statement_group
        coa_file = workbook_bytes(
            ["account_code", "name_ar", "name_en", "parent_code", "account_type", "statement_group", "is_cash", "active"],
            [[import_code, "حساب تحقق R4", "R4 verification account", parent_code, account_type, statement_group, False, True]],
        )
        files = {"file": ("coa-r4.xlsx", coa_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        match = ok(client.post("/api/v1/chart-of-accounts/import/validate?company_id=1", headers=admin, files=files))
        assert match["valid"] and match["summary"]["create"] == 1
        applied = ok(client.post("/api/v1/chart-of-accounts/import/apply?company_id=1", headers=admin, files={
            "file": ("coa-r4.xlsx", coa_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        }))
        assert applied["applied"]

        # 2) Opening balances validate, preserve their hash and post with maker-checker.
        with SessionLocal() as db:
            inventory_ids = {row.inventory_account_id for row in db.scalars(select(Item).where(Item.company_id == 1)).all()}
            asset = db.scalar(select(Account).where(
                Account.company_id == 1, Account.account_type == "ASSET",
                Account.is_postable.is_(True), Account.active.is_(True),
                Account.id.not_in(inventory_ids), Account.code.not_in(["112010", "211010"]),
            ).order_by(Account.code))
            equity = db.scalar(select(Account).where(
                Account.company_id == 1, Account.account_type == "EQUITY",
                Account.is_postable.is_(True), Account.active.is_(True),
            ).order_by(Account.code))
            assert asset and equity
            asset_code, equity_code = asset.code, equity.code
        opening_file = workbook_bytes(
            [
                "line_type", "account_code", "party_code", "item_code", "warehouse_code",
                "reference_code", "document_date", "due_date", "quantity", "unit_cost",
                "lot_number", "debit", "credit", "description",
            ],
            [
                ["GL", asset_code, None, None, None, None, None, None, None, None, None, 1234.56, 0, "Legacy asset"],
                ["GL", equity_code, None, None, None, None, None, None, None, None, None, 0, 1234.56, "Legacy equity"],
            ],
        )
        opening_files = {"file": ("opening-r4.xlsx", opening_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        validation = ok(client.post("/api/v1/opening-balances/validate?company_id=1", headers=admin, files=opening_files))
        assert validation["valid"] and validation["summary"]["difference"] == 0
        batch = ok(client.post(
            "/api/v1/opening-balances/import",
            params={"company_id": 1, "opening_date": "2026-01-01", "source_system": "LEGACY-R4"},
            headers=admin,
            files={"file": ("opening-r4.xlsx", opening_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ), 201)
        submitted = ok(client.post(f"/api/v1/opening-balances/{batch['id']}/submit", headers=admin))
        assert submitted["status"] == "PENDING_APPROVAL"
        assert client.post(f"/api/v1/opening-balances/{batch['id']}/approve-post", headers=admin).status_code == 409
        posted = ok(client.post(f"/api/v1/opening-balances/{batch['id']}/approve-post", headers=approver))
        assert posted["status"] == "POSTED" and posted["journal_id"]

        # 3) Item categories drive accounting and valuation; the type stays editable.
        with SessionLocal() as db:
            asset_account = db.scalar(select(Account).where(
                Account.company_id == 1, Account.account_type == "ASSET", Account.is_postable.is_(True),
            ).order_by(Account.code))
            expense_account = db.scalar(select(Account).where(
                Account.company_id == 1, Account.account_type == "EXPENSE", Account.is_postable.is_(True),
            ).order_by(Account.code))
            revenue_account = db.scalar(select(Account).where(
                Account.company_id == 1, Account.account_type == "REVENUE", Account.is_postable.is_(True),
            ).order_by(Account.code))
            assert asset_account and expense_account and revenue_account
            account_codes = (asset_account.code, expense_account.code, revenue_account.code)
        category = ok(client.post("/api/v1/inventory/item-categories", headers=admin, json={
            "company_id": 1, "code": "R4-CAT", "name_ar": "تصنيف تحقق", "name_en": "R4 Category",
            "default_item_type": "INVENTORY", "inventory_account_code": account_codes[0],
            "cogs_account_code": account_codes[1], "revenue_account_code": account_codes[2],
        }), 201)
        item = ok(client.post("/api/v1/inventory/items", headers=admin, json={
            "company_id": 1, "code": "R4-ITEM", "name_ar": "صنف تحقق", "name_en": "R4 Item",
            "item_type": "RAW_MATERIAL", "uom": "EA", "standard_cost": 10,
            "reorder_level": 2, "category_code": category["code"],
        }), 201)
        changed = ok(client.patch(f"/api/v1/inventory/items/{item['id']}", headers=admin, json={
            "company_id": 1, "item_type": "SERVICE", "category_code": "",
            "inventory_account_code": account_codes[0], "cogs_account_code": account_codes[1],
            "revenue_account_code": account_codes[2], "apply_category_defaults": False,
        }))
        assert changed["item_type"] == "SERVICE" and changed["category_code"] is None
        assert changed["valuation_method"] == "WEIGHTED_AVERAGE"

        # 4) Employee payroll linkage removes all hard blockers for that employee.
        branches = ok(client.get("/api/v1/enterprise/companies/1/branches", headers=admin))
        centers = ok(client.get("/api/v1/enterprise/companies/1/cost-centers", headers=admin))
        shifts = ok(client.get("/api/v1/hr/shifts?company_id=1", headers=admin))
        if not shifts:
            shift = ok(client.post("/api/v1/hr/shifts", headers=admin, json={
                "company_id": 1, "code": "R4-DAY", "name_ar": "وردية R4", "name_en": "R4 Day",
                "start_time": "08:00:00", "end_time": "17:00:00", "working_days": "6,0,1,2,3",
            }), 201)
        else:
            shift = shifts[0]
        employee = ok(client.post("/api/v1/payroll/employees", headers=admin, json={
            "company_id": 1, "employee_number": "R4-EMP", "name_ar": "موظف تحقق",
            "name_en": "R4 Employee", "nationality_group": "SAUDI", "hire_date": "2026-01-01",
            "basic_salary": 5000, "housing_allowance": 1250, "other_allowance": 0,
            "employee_gosi_rate": 9.75, "employer_gosi_rate": 11.75,
        }), 201)
        ok(client.post(f"/api/v1/payroll/employees/{employee['id']}/payroll-link", headers=admin, json={
            "company_id": 1, "branch_id": branches[0]["id"], "cost_center_id": centers[0]["id"],
            "shift_id": shift["id"], "effective_from": "2026-01-01",
            "iban": "SA0380000000608010167519", "salary_bank_code": "RJHI",
        }))
        readiness = ok(client.get("/api/v1/payroll/readiness?company_id=1&period_year=2026&period_month=7", headers=admin))
        employee_readiness = next(row for row in readiness["employees"] if row["employee_id"] == employee["id"])
        assert employee_readiness["ready"] and not employee_readiness["blockers"]

        with SessionLocal() as db:
            stored_batch = db.get(OpeningBalanceBatch, batch["id"])
            assert stored_batch.status == "POSTED"
            stored_item = db.get(Item, item["id"])
            assert stored_item.item_type == "SERVICE"
            stored_employee = db.get(Employee, employee["id"])
            assert stored_employee.branch_id and stored_employee.cost_center_id and stored_employee.iban
            assert not db.execute(select(JournalEntry.id).where(JournalEntry.total_debit != JournalEntry.total_credit)).all()
            assert not db.execute(text("PRAGMA foreign_key_check")).all()

    # 5) Navigation and AI drawer are structural release gates.
    navigation = (PROJECT_DIR / "frontend/src/dashboard/navigation.tsx").read_text(encoding="utf-8")
    routes = (PROJECT_DIR / "frontend/src/dashboard/routes.tsx").read_text(encoding="utf-8")
    ai_css = (PROJECT_DIR / "frontend/src/styles/rc27_4_ai_assistant_h5.css").read_text(encoding="utf-8")
    assistant = (PROJECT_DIR / "frontend/src/components/ai-assistant/CorvaxAiAssistant.tsx").read_text(encoding="utf-8")
    assert "salesReturns" in navigation and "purchaseReturns" in navigation
    assert "salesReturns" in routes and "purchaseReturns" in routes
    assert "openingBalances" in navigation and "openingBalances" in routes
    assert ".corvax-ai-backdrop {" not in ai_css and ".corvax-ai-open .dash .workspace" in ai_css
    assert 'aria-modal="false"' in assistant

    print("CORVAX RC27.4 R4 foundations: ALL SIX OWNER FINDINGS VERIFIED")
    DB_PATH.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
